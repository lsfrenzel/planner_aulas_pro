import json
import os
import logging
from functools import wraps
from flask import Flask, jsonify, request, render_template, send_file, Response, redirect, url_for, session, flash
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy.orm import DeclarativeBase
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.middleware.proxy_fix import ProxyFix
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from io import BytesIO

logging.basicConfig(level=logging.DEBUG)


class Base(DeclarativeBase):
    pass


db = SQLAlchemy(model_class=Base)

app = Flask(__name__)
app.secret_key = os.environ.get("SESSION_SECRET", "dev-secret-key-change-in-production")
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)

database_url = os.environ.get("DATABASE_URL")
if database_url and database_url.startswith("postgres://"):
    database_url = database_url.replace("postgres://", "postgresql://", 1)
    
if not database_url:
    raise RuntimeError("DATABASE_URL environment variable must be set. Please configure your PostgreSQL database connection.")
    
app.config["SQLALCHEMY_DATABASE_URI"] = database_url
app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
    "pool_recycle": 300,
    "pool_pre_ping": True,
}

db.init_app(app)

UPLOAD_FOLDER = os.path.join('static', 'uploads', 'profiles')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

DATA_FILE = "data/weeks.json"


def init_data():
    from models import User, Schedule, Turma
    
    admin = User.query.filter_by(role='admin').first()
    if not admin:
        admin = User(
            name="Administrador",
            email="admin@aula.com",
            password_hash=generate_password_hash("admin123"),
            role="admin",
            active=True
        )
        db.session.add(admin)
        db.session.commit()
        
        default_turma = Turma(
            user_id=admin.id,
            nome="Tecnico em Programacao de Jogos Digitais",
            descricao="Curso tecnico de desenvolvimento de jogos digitais",
            cor="blue",
            active=True
        )
        db.session.add(default_turma)
        db.session.commit()
        
        if os.path.exists(DATA_FILE):
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                weeks = json.load(f)
            for week in weeks:
                schedule = Schedule(
                    user_id=admin.id,
                    turma_id=default_turma.id,
                    semana=week.get("semana", 0),
                    atividades=week.get("atividades", ""),
                    unidade_curricular=week.get("unidadeCurricular", ""),
                    capacidades=week.get("capacidades", ""),
                    conhecimentos=week.get("conhecimentos", ""),
                    recursos=week.get("recursos", "")
                )
                db.session.add(schedule)
            db.session.commit()
    else:
        default_turma = Turma.query.filter_by(user_id=admin.id, nome="Tecnico em Programacao de Jogos Digitais").first()
        if not default_turma:
            default_turma = Turma(
                user_id=admin.id,
                nome="Tecnico em Programacao de Jogos Digitais",
                descricao="Curso tecnico de desenvolvimento de jogos digitais",
                cor="blue",
                active=True
            )
            db.session.add(default_turma)
            db.session.commit()
            
            orphan_schedules = Schedule.query.filter_by(user_id=admin.id, turma_id=None).all()
            for schedule in orphan_schedules:
                schedule.turma_id = default_turma.id
            db.session.commit()


def run_migrations():
    """Run database migrations for new columns if they don't exist"""
    from sqlalchemy import text
    
    migrations = [
        ("turmas", "carga_horaria", "ALTER TABLE turmas ADD COLUMN carga_horaria INTEGER"),
        ("turmas", "dias_aula", "ALTER TABLE turmas ADD COLUMN dias_aula VARCHAR(100)"),
        ("turmas", "horario_inicio", "ALTER TABLE turmas ADD COLUMN horario_inicio VARCHAR(10)"),
        ("turmas", "horario_fim", "ALTER TABLE turmas ADD COLUMN horario_fim VARCHAR(10)"),
        ("turmas", "data_inicio", "ALTER TABLE turmas ADD COLUMN data_inicio DATE"),
        ("turmas", "data_fim", "ALTER TABLE turmas ADD COLUMN data_fim DATE"),
        ("users", "cargo", "ALTER TABLE users ADD COLUMN cargo VARCHAR(100) DEFAULT ''"),
        ("users", "photo", "ALTER TABLE users ADD COLUMN photo VARCHAR(255) DEFAULT ''"),
        ("users", "photo_data", "ALTER TABLE users ADD COLUMN photo_data TEXT DEFAULT ''"),
        ("users", "photo_mimetype", "ALTER TABLE users ADD COLUMN photo_mimetype VARCHAR(50) DEFAULT ''"),
        ("schedules", "capacidades_completed", "ALTER TABLE schedules ADD COLUMN capacidades_completed TEXT DEFAULT ''"),
    ]
    
    for table, column, sql in migrations:
        try:
            check_sql = text(f"""
                SELECT column_name FROM information_schema.columns 
                WHERE table_name = :table AND column_name = :column
            """)
            result = db.session.execute(check_sql, {"table": table, "column": column}).fetchone()
            if not result:
                db.session.execute(text(sql))
                db.session.commit()
                logging.info(f"Migration: Added column {column} to {table}")
        except Exception as e:
            db.session.rollback()
            logging.warning(f"Migration warning for {table}.{column}: {e}")


with app.app_context():
    import models
    db.create_all()
    run_migrations()
    init_data()


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.path.startswith('/api/'):
                return jsonify({"error": "Nao autorizado"}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.path.startswith('/api/'):
                return jsonify({"error": "Nao autorizado"}), 401
            return redirect(url_for('login'))
        if session.get('user_role') != 'admin':
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.path.startswith('/api/'):
                return jsonify({"error": "Acesso negado"}), 403
            flash("Acesso negado. Apenas administradores podem acessar esta area.", "error")
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function


@app.route("/login", methods=["GET", "POST"])
def login():
    from models import User
    
    if 'user_id' in session:
        return redirect(url_for('index'))
    
    if request.method == "POST":
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "")
        
        if not email or not password:
            flash("Por favor, preencha todos os campos.", "error")
            return render_template("login.html")
        
        user = User.query.filter_by(email=email, active=True).first()
        
        if user and check_password_hash(user.password_hash, password):
            session['user_id'] = user.id
            session['user_name'] = user.name
            session['user_email'] = user.email
            session['user_role'] = user.role
            flash(f"Bem-vindo, {user.name}!", "success")
            return redirect(url_for('index'))
        else:
            flash("Email ou senha incorretos.", "error")
    
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    flash("Voce saiu do sistema.", "info")
    return redirect(url_for('login'))


@app.route("/")
@login_required
def index():
    from models import User
    user_data = User.query.get(session['user_id'])
    has_photo = bool(user_data and user_data.photo_data)
    return render_template("index.html", user=session, user_id=session['user_id'], has_photo=has_photo)


@app.route("/dashboard")
@login_required
def dashboard():
    from models import Schedule, User
    
    user_id = session['user_id']
    schedules = Schedule.query.filter_by(user_id=user_id).order_by(Schedule.semana).all()
    
    weeks = [s.to_dict() for s in schedules]
    
    unidades = set()
    recursos_set = set()
    for s in schedules:
        if s.unidade_curricular:
            unidades.add(s.unidade_curricular)
        if s.recursos:
            for r in s.recursos.split(','):
                recursos_set.add(r.strip())
    
    user_data = User.query.get(user_id)
    has_photo = bool(user_data and user_data.photo_data)
    
    return render_template("dashboard.html", 
                          user=session, 
                          user_id=user_id,
                          weeks=weeks,
                          unidades=list(unidades),
                          recursos=list(recursos_set),
                          has_photo=has_photo)


@app.route("/admin")
@admin_required
def admin_panel():
    return render_template("admin.html", user=session)


@app.route("/turmas")
@login_required
def turmas_page():
    from models import User
    user_data = User.query.get(session['user_id'])
    has_photo = bool(user_data and user_data.photo_data)
    return render_template("turmas.html", user=session, user_id=session['user_id'], has_photo=has_photo)


@app.route("/perfil")
@login_required
def perfil_page():
    from models import User
    user = User.query.get(session['user_id'])
    return render_template("perfil.html", user=session, user_data=user)


@app.route("/perfil/atualizar", methods=["POST"])
@login_required
def atualizar_perfil():
    from models import User
    import base64
    
    user = User.query.get(session['user_id'])
    if not user:
        flash("Usuario nao encontrado.", "error")
        return redirect(url_for('perfil_page'))
    
    name = request.form.get("name", "").strip()
    cargo = request.form.get("cargo", "").strip()
    
    if name:
        user.name = name
        session['user_name'] = name
    
    user.cargo = cargo
    
    if 'photo' in request.files:
        photo = request.files['photo']
        if photo and photo.filename and photo.filename.strip():
            allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
            ext = photo.filename.rsplit('.', 1)[-1].lower() if '.' in photo.filename else ''
            if ext in allowed_extensions:
                mime_types = {
                    'png': 'image/png',
                    'jpg': 'image/jpeg',
                    'jpeg': 'image/jpeg',
                    'gif': 'image/gif',
                    'webp': 'image/webp'
                }
                photo_data = base64.b64encode(photo.read()).decode('utf-8')
                user.photo_data = photo_data
                user.photo_mimetype = mime_types.get(ext, 'image/jpeg')
                user.photo = f"db_photo_{user.id}"
            else:
                flash("Formato de imagem nao permitido. Use PNG, JPG, JPEG, GIF ou WEBP.", "error")
                return redirect(url_for('perfil_page'))
    
    try:
        db.session.commit()
        flash("Perfil atualizado com sucesso!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao salvar perfil: {str(e)}", "error")
    
    return redirect(url_for('perfil_page'))


@app.route("/api/user-photo/<int:user_id>")
def get_user_photo(user_id):
    from models import User
    import base64
    
    user = User.query.get(user_id)
    if user and user.photo_data:
        try:
            photo_bytes = base64.b64decode(user.photo_data)
            return Response(photo_bytes, mimetype=user.photo_mimetype or 'image/jpeg')
        except Exception:
            pass
    
    svg_placeholder = '''<svg xmlns="http://www.w3.org/2000/svg" width="100" height="100" viewBox="0 0 100 100">
        <circle cx="50" cy="50" r="50" fill="#e2e8f0"/>
        <circle cx="50" cy="35" r="18" fill="#94a3b8"/>
        <ellipse cx="50" cy="80" rx="30" ry="25" fill="#94a3b8"/>
    </svg>'''
    return Response(svg_placeholder, mimetype='image/svg+xml')


@app.route("/perfil/alterar-senha", methods=["POST"])
@login_required
def alterar_senha():
    from models import User
    
    user = User.query.get(session['user_id'])
    if not user:
        flash("Usuario nao encontrado.", "error")
        return redirect(url_for('perfil_page'))
    
    senha_atual = request.form.get("senha_atual", "")
    nova_senha = request.form.get("nova_senha", "")
    confirmar_senha = request.form.get("confirmar_senha", "")
    
    if not check_password_hash(user.password_hash, senha_atual):
        flash("Senha atual incorreta.", "error")
        return redirect(url_for('perfil_page'))
    
    if len(nova_senha) < 6:
        flash("A nova senha deve ter pelo menos 6 caracteres.", "error")
        return redirect(url_for('perfil_page'))
    
    if nova_senha != confirmar_senha:
        flash("As senhas nao conferem.", "error")
        return redirect(url_for('perfil_page'))
    
    user.password_hash = generate_password_hash(nova_senha)
    db.session.commit()
    
    flash("Senha alterada com sucesso!", "success")
    return redirect(url_for('perfil_page'))


@app.route("/api/turmas", methods=["GET"])
@login_required
def get_turmas():
    from models import Turma
    user_id = session['user_id']
    turmas = Turma.query.filter_by(user_id=user_id, active=True).order_by(Turma.nome).all()
    return jsonify([t.to_dict() for t in turmas])


@app.route("/api/turmas", methods=["POST"])
@login_required
def add_turma():
    from models import Turma
    from datetime import datetime
    
    user_id = session['user_id']
    data = request.get_json()
    
    if not data:
        return jsonify({"error": "Dados invalidos"}), 400
    
    nome = data.get("nome", "").strip()
    descricao = data.get("descricao", "").strip()
    cor = data.get("cor", "blue")
    carga_horaria = data.get("carga_horaria", 0)
    dias_aula = data.get("dias_aula", "")
    horario_inicio = data.get("horario_inicio", "")
    horario_fim = data.get("horario_fim", "")
    data_inicio_str = data.get("data_inicio", "")
    data_fim_str = data.get("data_fim", "")
    
    if not nome:
        return jsonify({"error": "Nome da turma e obrigatorio"}), 400
    
    data_inicio = None
    data_fim = None
    if data_inicio_str:
        try:
            data_inicio = datetime.strptime(data_inicio_str, "%Y-%m-%d").date()
        except ValueError:
            pass
    if data_fim_str:
        try:
            data_fim = datetime.strptime(data_fim_str, "%Y-%m-%d").date()
        except ValueError:
            pass
    
    turma = Turma(
        user_id=user_id,
        nome=nome,
        descricao=descricao,
        cor=cor,
        carga_horaria=carga_horaria,
        dias_aula=dias_aula,
        horario_inicio=horario_inicio,
        horario_fim=horario_fim,
        data_inicio=data_inicio,
        data_fim=data_fim,
        active=True
    )
    db.session.add(turma)
    db.session.commit()
    
    return jsonify(turma.to_dict()), 201


@app.route("/api/turmas/<int:turma_id>", methods=["PUT"])
@login_required
def update_turma(turma_id):
    from models import Turma
    from datetime import datetime
    
    user_id = session['user_id']
    data = request.get_json()
    
    if not data:
        return jsonify({"error": "Dados invalidos"}), 400
    
    turma = Turma.query.filter_by(id=turma_id, user_id=user_id).first()
    
    if not turma:
        return jsonify({"error": "Turma nao encontrada"}), 404
    
    turma.nome = data.get("nome", turma.nome).strip()
    turma.descricao = data.get("descricao", turma.descricao).strip()
    turma.cor = data.get("cor", turma.cor)
    turma.carga_horaria = data.get("carga_horaria", turma.carga_horaria)
    turma.dias_aula = data.get("dias_aula", turma.dias_aula)
    turma.horario_inicio = data.get("horario_inicio", turma.horario_inicio)
    turma.horario_fim = data.get("horario_fim", turma.horario_fim)
    
    data_inicio_str = data.get("data_inicio", "")
    data_fim_str = data.get("data_fim", "")
    
    if data_inicio_str:
        try:
            turma.data_inicio = datetime.strptime(data_inicio_str, "%Y-%m-%d").date()
        except ValueError:
            pass
    elif "data_inicio" in data and not data_inicio_str:
        turma.data_inicio = None
        
    if data_fim_str:
        try:
            turma.data_fim = datetime.strptime(data_fim_str, "%Y-%m-%d").date()
        except ValueError:
            pass
    elif "data_fim" in data and not data_fim_str:
        turma.data_fim = None
    
    db.session.commit()
    
    return jsonify(turma.to_dict())


@app.route("/api/turmas/<int:turma_id>", methods=["DELETE"])
@login_required
def delete_turma(turma_id):
    from models import Turma
    
    user_id = session['user_id']
    turma = Turma.query.filter_by(id=turma_id, user_id=user_id).first()
    
    if not turma:
        return jsonify({"error": "Turma nao encontrada"}), 404
    
    turma.active = False
    db.session.commit()
    
    return jsonify({"message": "Turma excluida com sucesso"})


@app.route("/api/users", methods=["GET"])
@admin_required
def get_users():
    from models import User
    users = User.query.order_by(User.id).all()
    return jsonify([u.to_dict() for u in users])


@app.route("/api/users", methods=["POST"])
@admin_required
def add_user():
    from models import User
    
    data = request.get_json()
    
    if not data:
        return jsonify({"error": "Dados invalidos"}), 400
    
    name = data.get("name", "").strip()
    email = data.get("email", "").strip()
    password = data.get("password", "")
    role = data.get("role", "user")
    
    if not name or not email or not password:
        return jsonify({"error": "Nome, email e senha sao obrigatorios"}), 400
    
    if role not in ["user", "admin"]:
        role = "user"
    
    existing = User.query.filter_by(email=email).first()
    if existing:
        return jsonify({"error": "Email ja cadastrado"}), 400
    
    user = User(
        name=name,
        email=email,
        password_hash=generate_password_hash(password),
        role=role,
        active=True
    )
    db.session.add(user)
    db.session.commit()
    
    return jsonify(user.to_dict()), 201


@app.route("/api/users/<int:user_id>", methods=["PUT"])
@admin_required
def update_user(user_id):
    from models import User
    
    data = request.get_json()
    
    if not data:
        return jsonify({"error": "Dados invalidos"}), 400
    
    user = User.query.get(user_id)
    
    if not user:
        return jsonify({"error": "Usuario nao encontrado"}), 404
    
    user.name = data.get("name", user.name).strip()
    email = data.get("email", user.email).strip()
    role = data.get("role", user.role)
    active = data.get("active", user.active)
    
    if role not in ["user", "admin"]:
        role = user.role
    
    if email != user.email:
        existing = User.query.filter_by(email=email).first()
        if existing:
            return jsonify({"error": "Email ja cadastrado"}), 400
        user.email = email
    
    user.role = role
    user.active = active
    
    if "password" in data and data["password"]:
        user.password_hash = generate_password_hash(data["password"])
    
    db.session.commit()
    
    return jsonify(user.to_dict())


@app.route("/api/users/<int:user_id>", methods=["DELETE"])
@admin_required
def delete_user(user_id):
    from models import User
    
    if session.get('user_id') == user_id:
        return jsonify({"error": "Voce nao pode excluir sua propria conta"}), 400
    
    user = User.query.get(user_id)
    
    if not user:
        return jsonify({"error": "Usuario nao encontrado"}), 404
    
    db.session.delete(user)
    db.session.commit()
    
    return jsonify({"message": "Usuario excluido com sucesso"})


@app.route("/api/weeks", methods=["GET"])
@login_required
def get_weeks():
    from models import Schedule
    
    user_id = session['user_id']
    turma_id = request.args.get('turma_id', type=int)
    
    query = Schedule.query.filter_by(user_id=user_id)
    if turma_id:
        query = query.filter_by(turma_id=turma_id)
    
    schedules = query.order_by(Schedule.semana).all()
    return jsonify([s.to_dict() for s in schedules])


@app.route("/api/weeks/<int:week_id>", methods=["GET"])
@login_required
def get_week(week_id):
    from models import Schedule
    
    user_id = session['user_id']
    schedule = Schedule.query.filter_by(id=week_id, user_id=user_id).first()
    
    if schedule:
        return jsonify(schedule.to_dict())
    return jsonify({"error": "Semana nao encontrada"}), 404


@app.route("/api/weeks", methods=["POST"])
@login_required
def add_week():
    from models import Schedule, Turma
    
    user_id = session['user_id']
    data = request.get_json()
    
    if not data:
        return jsonify({"error": "Dados invalidos"}), 400
    
    turma_id = data.get("turma_id")
    if not turma_id:
        return jsonify({"error": "Turma e obrigatoria. Selecione uma turma antes de adicionar semanas."}), 400
    
    turma = Turma.query.filter_by(id=turma_id, user_id=user_id, active=True).first()
    if not turma:
        return jsonify({"error": "Turma nao encontrada"}), 404
    
    max_semana = db.session.query(db.func.max(Schedule.semana)).filter_by(user_id=user_id, turma_id=turma_id).scalar() or 0
    
    schedule = Schedule(
        user_id=user_id,
        turma_id=turma_id,
        semana=data.get("semana", max_semana + 1),
        atividades=data.get("atividades", ""),
        unidade_curricular=data.get("unidadeCurricular", ""),
        capacidades=data.get("capacidades", ""),
        conhecimentos=data.get("conhecimentos", ""),
        recursos=data.get("recursos", "")
    )
    
    db.session.add(schedule)
    db.session.commit()
    
    return jsonify(schedule.to_dict()), 201


@app.route("/api/weeks/<int:week_id>", methods=["PUT"])
@login_required
def update_week(week_id):
    from models import Schedule
    
    user_id = session['user_id']
    data = request.get_json()
    
    if not data:
        return jsonify({"error": "Dados invalidos"}), 400
    
    schedule = Schedule.query.filter_by(id=week_id, user_id=user_id).first()
    
    if not schedule:
        return jsonify({"error": "Semana nao encontrada"}), 404
    
    schedule.atividades = data.get("atividades", schedule.atividades)
    schedule.unidade_curricular = data.get("unidadeCurricular", schedule.unidade_curricular)
    schedule.capacidades = data.get("capacidades", schedule.capacidades)
    schedule.conhecimentos = data.get("conhecimentos", schedule.conhecimentos)
    schedule.recursos = data.get("recursos", schedule.recursos)
    
    db.session.commit()
    
    return jsonify(schedule.to_dict())


@app.route("/api/weeks/<int:week_id>", methods=["DELETE"])
@login_required
def delete_week(week_id):
    from models import Schedule
    
    user_id = session['user_id']
    schedule = Schedule.query.filter_by(id=week_id, user_id=user_id).first()
    
    if not schedule:
        return jsonify({"error": "Semana nao encontrada"}), 404
    
    db.session.delete(schedule)
    db.session.commit()
    
    return jsonify({"message": "Semana excluida com sucesso"})


@app.route("/api/weeks/<int:week_id>/toggle-complete", methods=["POST"])
@login_required
def toggle_week_complete(week_id):
    from models import Schedule
    
    user_id = session['user_id']
    schedule = Schedule.query.filter_by(id=week_id, user_id=user_id).first()
    
    if not schedule:
        return jsonify({"error": "Semana nao encontrada"}), 404
    
    schedule.completed = not schedule.completed
    db.session.commit()
    
    return jsonify(schedule.to_dict())


@app.route("/api/weeks/<int:week_id>/toggle-capacidade", methods=["POST"])
@login_required
def toggle_capacidade(week_id):
    from models import Schedule
    
    user_id = session['user_id']
    schedule = Schedule.query.filter_by(id=week_id, user_id=user_id).first()
    
    if not schedule:
        return jsonify({"error": "Semana nao encontrada"}), 404
    
    data = request.get_json()
    capacidade_index = data.get("index")
    
    if capacidade_index is None:
        return jsonify({"error": "Indice da capacidade nao informado"}), 400
    
    completed_list = schedule.capacidades_completed.split(',') if schedule.capacidades_completed else []
    completed_list = [x for x in completed_list if x]
    
    index_str = str(capacidade_index)
    if index_str in completed_list:
        completed_list.remove(index_str)
    else:
        completed_list.append(index_str)
    
    schedule.capacidades_completed = ','.join(completed_list)
    db.session.commit()
    
    return jsonify({
        "id": schedule.id,
        "capacidades_completed": schedule.capacidades_completed
    })


@app.route("/api/turmas/progress", methods=["GET"])
@login_required
def get_turmas_progress():
    from models import Turma, Schedule
    
    user_id = session['user_id']
    turmas = Turma.query.filter_by(user_id=user_id, active=True).all()
    
    progress_data = []
    for turma in turmas:
        schedules = Schedule.query.filter_by(user_id=user_id, turma_id=turma.id).all()
        
        total_weeks = len(schedules)
        completed_weeks = sum(1 for s in schedules if s.completed)
        
        total_capacidades = 0
        completed_capacidades = 0
        
        for s in schedules:
            caps = [c.strip() for c in s.capacidades.split('\n') if c.strip()]
            total_capacidades += len(caps)
            
            completed_list = s.capacidades_completed.split(',') if s.capacidades_completed else []
            completed_list = [x for x in completed_list if x]
            completed_capacidades += len(completed_list)
        
        weeks_percent = round((completed_weeks / total_weeks * 100) if total_weeks > 0 else 0)
        caps_percent = round((completed_capacidades / total_capacidades * 100) if total_capacidades > 0 else 0)
        
        progress_data.append({
            'id': turma.id,
            'nome': turma.nome,
            'cor': turma.cor,
            'total_weeks': total_weeks,
            'completed_weeks': completed_weeks,
            'weeks_percent': weeks_percent,
            'total_capacidades': total_capacidades,
            'completed_capacidades': completed_capacidades,
            'capacidades_percent': caps_percent
        })
    
    return jsonify(progress_data)


@app.route("/api/migrate")
def run_migration():
    from models import User, Schedule, Turma
    
    try:
        db.session.execute(db.text("""
            CREATE TABLE IF NOT EXISTS turmas (
                id SERIAL PRIMARY KEY,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                nome VARCHAR(200) NOT NULL,
                descricao TEXT DEFAULT '',
                cor VARCHAR(20) DEFAULT 'blue',
                active BOOLEAN DEFAULT TRUE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """))
        db.session.execute(db.text("ALTER TABLE schedules ADD COLUMN IF NOT EXISTS completed BOOLEAN DEFAULT FALSE;"))
        db.session.execute(db.text("ALTER TABLE schedules ADD COLUMN IF NOT EXISTS turma_id INTEGER REFERENCES turmas(id) ON DELETE CASCADE;"))
        db.session.commit()
        
        users = User.query.all()
        for user in users:
            orphan_schedules = db.session.execute(
                db.text("SELECT id FROM schedules WHERE user_id = :uid AND turma_id IS NULL"),
                {"uid": user.id}
            ).fetchall()
            
            if orphan_schedules:
                default_turma = Turma.query.filter_by(user_id=user.id).first()
                if not default_turma:
                    default_turma = Turma(
                        user_id=user.id,
                        nome="Tecnico em Programacao de Jogos Digitais",
                        descricao="Turma padrao criada automaticamente",
                        cor="blue",
                        active=True
                    )
                    db.session.add(default_turma)
                    db.session.commit()
                
                db.session.execute(
                    db.text("UPDATE schedules SET turma_id = :tid WHERE user_id = :uid AND turma_id IS NULL"),
                    {"tid": default_turma.id, "uid": user.id}
                )
                db.session.commit()
        
        try:
            db.session.execute(db.text("ALTER TABLE schedules ALTER COLUMN turma_id SET NOT NULL;"))
            db.session.commit()
        except Exception:
            db.session.rollback()
        
        return jsonify({"success": True, "message": "Migracao executada com sucesso! Schedules associados a turmas, constraint NOT NULL aplicada."})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/export/json")
@login_required
def export_json():
    from models import Schedule
    
    user_id = session['user_id']
    schedules = Schedule.query.filter_by(user_id=user_id).order_by(Schedule.semana).all()
    weeks = [s.to_dict() for s in schedules]
    
    response = Response(
        json.dumps(weeks, ensure_ascii=False, indent=2),
        mimetype="application/json",
        headers={"Content-Disposition": "attachment;filename=cronograma.json"}
    )
    return response


@app.route("/api/export/pdf")
@login_required
def export_pdf():
    from models import Schedule, Turma
    from datetime import datetime
    
    user_id = session['user_id']
    turma_id = request.args.get('turma_id', type=int)
    
    query = Schedule.query.filter_by(user_id=user_id)
    if turma_id:
        query = query.filter_by(turma_id=turma_id)
    
    schedules = query.order_by(Schedule.semana).all()
    weeks = [s.to_dict() for s in schedules]
    
    turma = None
    if turma_id:
        turma = Turma.query.filter_by(id=turma_id, user_id=user_id).first()
    
    all_capacidades_desenvolvidas = []
    total_capacidades = 0
    total_completed = 0
    completed_weeks = 0
    
    for week in weeks:
        if week.get('completed'):
            completed_weeks += 1
        
        caps = [c.strip() for c in week.get('capacidades', '').split('\n') if c.strip()]
        completed_list = week.get('capacidades_completed', '').split(',') if week.get('capacidades_completed') else []
        completed_list = [x for x in completed_list if x]
        
        total_capacidades += len(caps)
        total_completed += len(completed_list)
        
        for idx, cap in enumerate(caps):
            is_completed = str(idx) in completed_list
            if is_completed:
                all_capacidades_desenvolvidas.append({
                    'semana': week['semana'],
                    'capacidade': cap,
                    'unidade': week.get('unidadeCurricular', '')
                })
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1*cm,
        leftMargin=1*cm,
        topMargin=1*cm,
        bottomMargin=1*cm
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=10,
        alignment=1
    )
    
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=12,
        spaceAfter=5,
        alignment=1,
        textColor=colors.HexColor('#4B5563')
    )
    
    info_style = ParagraphStyle(
        'InfoStyle',
        parent=styles['Normal'],
        fontSize=9,
        spaceAfter=3,
        textColor=colors.HexColor('#374151')
    )
    
    cell_style = ParagraphStyle(
        'CellStyle',
        parent=styles['Normal'],
        fontSize=7,
        leading=9
    )
    
    cell_completed_style = ParagraphStyle(
        'CellCompletedStyle',
        parent=styles['Normal'],
        fontSize=7,
        leading=9,
        textColor=colors.HexColor('#059669')
    )
    
    section_title_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceBefore=20,
        spaceAfter=10,
        textColor=colors.HexColor('#059669')
    )
    
    if turma:
        elements.append(Paragraph(f"Cronograma - {turma.nome}", title_style))
        if turma.descricao:
            elements.append(Paragraph(turma.descricao, subtitle_style))
        elements.append(Spacer(1, 10))
        
        info_items = []
        if turma.carga_horaria:
            info_items.append(f"<b>Carga Horaria:</b> {turma.carga_horaria}h")
        if turma.dias_aula:
            info_items.append(f"<b>Dias de Aula:</b> {turma.dias_aula}")
        if turma.horario_inicio and turma.horario_fim:
            info_items.append(f"<b>Horario:</b> {turma.horario_inicio} - {turma.horario_fim}")
        elif turma.horario_inicio:
            info_items.append(f"<b>Horario:</b> {turma.horario_inicio}")
        
        if turma.data_inicio or turma.data_fim:
            data_inicio_str = turma.data_inicio.strftime('%d/%m/%Y') if turma.data_inicio else '-'
            data_fim_str = turma.data_fim.strftime('%d/%m/%Y') if turma.data_fim else '-'
            info_items.append(f"<b>Periodo:</b> {data_inicio_str} a {data_fim_str}")
        
        if info_items:
            info_text = " &nbsp;&nbsp;|&nbsp;&nbsp; ".join(info_items)
            elements.append(Paragraph(info_text, info_style))
            elements.append(Spacer(1, 10))
    else:
        elements.append(Paragraph("Aula Planner Pro - Cronograma Completo", title_style))
        elements.append(Spacer(1, 10))
    
    progress_percent_weeks = round((completed_weeks / len(weeks) * 100) if len(weeks) > 0 else 0)
    progress_percent_caps = round((total_completed / total_capacidades * 100) if total_capacidades > 0 else 0)
    
    progress_text = f"<b>Progresso:</b> {completed_weeks}/{len(weeks)} semanas concluidas ({progress_percent_weeks}%) | {total_completed}/{total_capacidades} capacidades desenvolvidas ({progress_percent_caps}%)"
    elements.append(Paragraph(progress_text, info_style))
    elements.append(Spacer(1, 15))
    
    headers = ["Status", "Semana", "Atividades", "Unidade Curricular", "Capacidades", "Conhecimentos", "Recursos"]
    
    if not turma_id:
        headers.insert(2, "Turma")
    
    data = [headers]
    for week in weeks:
        status = "Concluida" if week.get('completed') else "Pendente"
        
        caps = [c.strip() for c in week.get('capacidades', '').split('\n') if c.strip()]
        completed_list = week.get('capacidades_completed', '').split(',') if week.get('capacidades_completed') else []
        completed_list = [x for x in completed_list if x]
        
        capacidades_formatted = []
        for idx, cap in enumerate(caps):
            if str(idx) in completed_list:
                capacidades_formatted.append(f"[OK] {cap}")
            else:
                capacidades_formatted.append(f"[ ] {cap}")
        
        capacidades_text = "\n".join(capacidades_formatted) if capacidades_formatted else week.get('capacidades', '')
        
        if turma_id:
            row = [
                status,
                str(week["semana"]),
                Paragraph(week["atividades"], cell_style),
                Paragraph(week["unidadeCurricular"], cell_style),
                Paragraph(capacidades_text.replace('\n', '<br/>'), cell_completed_style if week.get('completed') else cell_style),
                Paragraph(week["conhecimentos"], cell_style),
                Paragraph(week["recursos"], cell_style)
            ]
        else:
            row = [
                status,
                str(week["semana"]),
                Paragraph(week.get("turma_nome", ""), cell_style),
                Paragraph(week["atividades"], cell_style),
                Paragraph(week["unidadeCurricular"], cell_style),
                Paragraph(capacidades_text.replace('\n', '<br/>'), cell_completed_style if week.get('completed') else cell_style),
                Paragraph(week["conhecimentos"], cell_style),
                Paragraph(week["recursos"], cell_style)
            ]
        data.append(row)
    
    if turma_id:
        col_widths = [1.5*cm, 1.2*cm, 4.5*cm, 3.5*cm, 5.5*cm, 4*cm, 3.5*cm]
    else:
        col_widths = [1.3*cm, 1*cm, 2.5*cm, 4*cm, 3*cm, 5*cm, 3.5*cm, 3*cm]
    
    table = Table(data, colWidths=col_widths, repeatRows=1)
    
    table_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
    ]
    
    for idx, week in enumerate(weeks, 1):
        if week.get('completed'):
            table_style.append(('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#D1FAE5')))
            table_style.append(('TEXTCOLOR', (0, idx), (0, idx), colors.HexColor('#059669')))
    
    table.setStyle(TableStyle(table_style))
    elements.append(table)
    
    if all_capacidades_desenvolvidas:
        elements.append(Spacer(1, 30))
        elements.append(Paragraph("Capacidades Desenvolvidas", section_title_style))
        elements.append(Spacer(1, 10))
        
        summary_info = f"Total de {len(all_capacidades_desenvolvidas)} capacidades desenvolvidas ao longo do curso."
        elements.append(Paragraph(summary_info, info_style))
        elements.append(Spacer(1, 10))
        
        caps_headers = ["Semana", "Unidade Curricular", "Capacidade Desenvolvida"]
        caps_data = [caps_headers]
        
        for cap_info in all_capacidades_desenvolvidas:
            caps_data.append([
                str(cap_info['semana']),
                Paragraph(cap_info['unidade'], cell_style),
                Paragraph(cap_info['capacidade'], cell_style)
            ])
        
        caps_col_widths = [2*cm, 6*cm, 18*cm]
        caps_table = Table(caps_data, colWidths=caps_col_widths, repeatRows=1)
        caps_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#ECFDF5')]),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ]))
        elements.append(caps_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"cronograma_{turma.nome.replace(' ', '_')}.pdf" if turma else "cronograma.pdf"
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )


@app.route("/api/export/xlsx")
@login_required
def export_xlsx():
    from models import Schedule, Turma
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    user_id = session['user_id']
    turma_id = request.args.get('turma_id', type=int)
    
    query = Schedule.query.filter_by(user_id=user_id)
    if turma_id:
        query = query.filter_by(turma_id=turma_id)
    
    schedules = query.order_by(Schedule.semana).all()
    weeks = [s.to_dict() for s in schedules]
    
    turma = None
    if turma_id:
        turma = Turma.query.filter_by(id=turma_id, user_id=user_id).first()
    
    all_capacidades_desenvolvidas = []
    total_capacidades = 0
    total_completed = 0
    completed_weeks = 0
    
    for week in weeks:
        if week.get('completed'):
            completed_weeks += 1
        
        caps = [c.strip() for c in week.get('capacidades', '').split('\n') if c.strip()]
        completed_list = week.get('capacidades_completed', '').split(',') if week.get('capacidades_completed') else []
        completed_list = [x for x in completed_list if x]
        
        total_capacidades += len(caps)
        total_completed += len(completed_list)
        
        for idx, cap in enumerate(caps):
            is_completed = str(idx) in completed_list
            if is_completed:
                all_capacidades_desenvolvidas.append({
                    'semana': week['semana'],
                    'capacidade': cap,
                    'unidade': week.get('unidadeCurricular', '')
                })
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Cronograma"
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    
    completed_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    completed_font = Font(color="059669")
    green_header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    
    current_row = 1
    
    if turma:
        ws.merge_cells(f'A{current_row}:H{current_row}')
        title_cell = ws.cell(row=current_row, column=1, value=f"Cronograma - {turma.nome}")
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal="center")
        current_row += 1
        
        if turma.descricao:
            ws.merge_cells(f'A{current_row}:H{current_row}')
            desc_cell = ws.cell(row=current_row, column=1, value=turma.descricao)
            desc_cell.alignment = Alignment(horizontal="center")
            current_row += 1
        
        info_parts = []
        if turma.carga_horaria:
            info_parts.append(f"Carga Horaria: {turma.carga_horaria}h")
        if turma.dias_aula:
            info_parts.append(f"Dias de Aula: {turma.dias_aula}")
        if turma.horario_inicio and turma.horario_fim:
            info_parts.append(f"Horario: {turma.horario_inicio} - {turma.horario_fim}")
        if turma.data_inicio or turma.data_fim:
            data_inicio_str = turma.data_inicio.strftime('%d/%m/%Y') if turma.data_inicio else '-'
            data_fim_str = turma.data_fim.strftime('%d/%m/%Y') if turma.data_fim else '-'
            info_parts.append(f"Periodo: {data_inicio_str} a {data_fim_str}")
        
        if info_parts:
            ws.merge_cells(f'A{current_row}:H{current_row}')
            info_cell = ws.cell(row=current_row, column=1, value=" | ".join(info_parts))
            info_cell.alignment = Alignment(horizontal="center")
            current_row += 1
        
        current_row += 1
    else:
        ws.merge_cells(f'A{current_row}:H{current_row}')
        title_cell = ws.cell(row=current_row, column=1, value="Aula Planner Pro - Cronograma Completo")
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal="center")
        current_row += 2
    
    progress_percent_weeks = round((completed_weeks / len(weeks) * 100) if len(weeks) > 0 else 0)
    progress_percent_caps = round((total_completed / total_capacidades * 100) if total_capacidades > 0 else 0)
    
    ws.merge_cells(f'A{current_row}:H{current_row}')
    progress_cell = ws.cell(row=current_row, column=1, 
                            value=f"Progresso: {completed_weeks}/{len(weeks)} semanas concluidas ({progress_percent_weeks}%) | {total_completed}/{total_capacidades} capacidades desenvolvidas ({progress_percent_caps}%)")
    progress_cell.font = Font(bold=True, color="059669")
    progress_cell.alignment = Alignment(horizontal="center")
    current_row += 2
    
    if turma_id:
        headers = ["Status", "Semana", "Atividades", "Unidade Curricular", "Capacidades", "Conhecimentos", "Recursos"]
    else:
        headers = ["Status", "Semana", "Turma", "Atividades", "Unidade Curricular", "Capacidades", "Conhecimentos", "Recursos"]
    
    header_row = current_row
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    current_row += 1
    
    for week in weeks:
        status = "Concluida" if week.get('completed') else "Pendente"
        
        caps = [c.strip() for c in week.get('capacidades', '').split('\n') if c.strip()]
        completed_list = week.get('capacidades_completed', '').split(',') if week.get('capacidades_completed') else []
        completed_list = [x for x in completed_list if x]
        
        capacidades_formatted = []
        for idx, cap in enumerate(caps):
            if str(idx) in completed_list:
                capacidades_formatted.append(f"[OK] {cap}")
            else:
                capacidades_formatted.append(f"[ ] {cap}")
        
        capacidades_text = "\n".join(capacidades_formatted) if capacidades_formatted else week.get('capacidades', '')
        
        if turma_id:
            row_data = [
                status,
                week["semana"],
                week["atividades"],
                week["unidadeCurricular"],
                capacidades_text,
                week["conhecimentos"],
                week["recursos"]
            ]
        else:
            row_data = [
                status,
                week["semana"],
                week.get("turma_nome", ""),
                week["atividades"],
                week["unidadeCurricular"],
                capacidades_text,
                week["conhecimentos"],
                week["recursos"]
            ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.alignment = cell_alignment if col > 2 else center_alignment
            cell.border = thin_border
            
            if week.get('completed'):
                cell.fill = completed_fill
                if col == 1:
                    cell.font = completed_font
        
        current_row += 1
    
    if turma_id:
        col_widths = [12, 10, 40, 25, 40, 30, 25]
    else:
        col_widths = [12, 10, 20, 35, 22, 35, 25, 22]
    
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    if all_capacidades_desenvolvidas:
        ws_caps = wb.create_sheet(title="Capacidades Desenvolvidas")
        
        ws_caps.merge_cells('A1:C1')
        title_cell = ws_caps.cell(row=1, column=1, value="Capacidades Desenvolvidas")
        title_cell.font = Font(bold=True, size=16, color="059669")
        title_cell.alignment = Alignment(horizontal="center")
        
        ws_caps.merge_cells('A2:C2')
        summary_cell = ws_caps.cell(row=2, column=1, 
                                    value=f"Total de {len(all_capacidades_desenvolvidas)} capacidades desenvolvidas ao longo do curso")
        summary_cell.alignment = Alignment(horizontal="center")
        
        caps_headers = ["Semana", "Unidade Curricular", "Capacidade Desenvolvida"]
        for col, header in enumerate(caps_headers, 1):
            cell = ws_caps.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = green_header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        caps_row = 5
        for cap_info in all_capacidades_desenvolvidas:
            ws_caps.cell(row=caps_row, column=1, value=cap_info['semana']).alignment = center_alignment
            ws_caps.cell(row=caps_row, column=1).border = thin_border
            
            ws_caps.cell(row=caps_row, column=2, value=cap_info['unidade']).alignment = cell_alignment
            ws_caps.cell(row=caps_row, column=2).border = thin_border
            
            ws_caps.cell(row=caps_row, column=3, value=cap_info['capacidade']).alignment = cell_alignment
            ws_caps.cell(row=caps_row, column=3).border = thin_border
            
            if caps_row % 2 == 0:
                for col in range(1, 4):
                    ws_caps.cell(row=caps_row, column=col).fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
            
            caps_row += 1
        
        ws_caps.column_dimensions['A'].width = 10
        ws_caps.column_dimensions['B'].width = 30
        ws_caps.column_dimensions['C'].width = 80
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"cronograma_{turma.nome.replace(' ', '_')}.xlsx" if turma else "cronograma.xlsx"
    
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@app.after_request
def add_header(response):
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
